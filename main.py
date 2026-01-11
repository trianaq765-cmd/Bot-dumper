import discord,os,io,re,time,random,logging,sqlite3,base64,json
from collections import defaultdict
from dataclasses import dataclass
from urllib.parse import quote
from discord import app_commands
from discord.ext import commands
from keep_alive import keep_alive

logging.basicConfig(level=logging.WARNING,format='%(asctime)s|%(levelname)s|%(message)s')
logger=logging.getLogger(__name__)
logger.setLevel(logging.INFO)

DISCORD_TOKEN=os.getenv("DISCORD_TOKEN")
KEY_GROQ=os.getenv("GROQ_API_KEY")
KEY_GEMINI=os.getenv("GEMINI_API_KEY")
KEY_OPENAI=os.getenv("OPENAI_API_KEY")
SCRAPER_KEY=os.getenv("SCRAPER_API_KEY")
KEY_LUAOBF=os.getenv("LUAOBF_API_KEY")
OWNER_IDS=[int(x)for x in os.getenv("OWNER_IDS","0").split(",")if x.isdigit()]

if not DISCORD_TOKEN:print("❌ NO TOKEN!");exit(1)

intents=discord.Intents.default()
intents.message_content=True
bot=commands.Bot(command_prefix="!",intents=intents)

# ==============================================================================
# 📦 LAZY IMPORTS
# ==============================================================================
_groq=_openai=_genai=_curl=_requests=_aiohttp=_pd=_openpyxl=None

def get_groq():
    global _groq
    if _groq is None and KEY_GROQ:
        from groq import Groq
        _groq=Groq(api_key=KEY_GROQ)
    return _groq

def get_openai():
    global _openai
    if _openai is None and KEY_OPENAI:
        from openai import OpenAI
        _openai=OpenAI(api_key=KEY_OPENAI)
    return _openai

def get_genai():
    global _genai
    if _genai is None and KEY_GEMINI:
        import google.generativeai as genai
        genai.configure(api_key=KEY_GEMINI)
        _genai=genai
    return _genai

def get_curl():
    global _curl
    if _curl is None:
        from curl_cffi import requests as curl_requests
        _curl=curl_requests
    return _curl

def get_requests():
    global _requests
    if _requests is None:
        import requests
        _requests=requests
    return _requests

async def get_aiohttp():
    global _aiohttp
    if _aiohttp is None:
        import aiohttp
        _aiohttp=aiohttp
    return _aiohttp

def get_pandas():
    global _pd
    if _pd is None:
        import pandas as pd
        _pd=pd
    return _pd

def get_openpyxl():
    global _openpyxl
    if _openpyxl is None:
        import openpyxl
        _openpyxl=openpyxl
    return _openpyxl

# ==============================================================================
# 💾 DATABASE
# ==============================================================================
class Database:
    def __init__(self,path="bot.db"):
        self.conn=sqlite3.connect(path,check_same_thread=False)
        self.conn.executescript('''
            CREATE TABLE IF NOT EXISTS history(id INTEGER PRIMARY KEY,uid INTEGER,gid INTEGER,cmd TEXT,prompt TEXT,resp TEXT,model TEXT,ts TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
            CREATE TABLE IF NOT EXISTS blacklist(uid INTEGER PRIMARY KEY,reason TEXT,by INTEGER);
            CREATE TABLE IF NOT EXISTS stats(id INTEGER PRIMARY KEY,cmd TEXT,uid INTEGER,gid INTEGER,ok INTEGER,t REAL,ts TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
        ''')
    def log(self,uid,gid,cmd,p,r,m):self.conn.execute('INSERT INTO history(uid,gid,cmd,prompt,resp,model)VALUES(?,?,?,?,?,?)',(uid,gid,cmd,p,r[:4000],m));self.conn.commit()
    def hist(self,uid,n=5):return self.conn.execute('SELECT prompt,resp,ts FROM history WHERE uid=? ORDER BY ts DESC LIMIT ?',(uid,n)).fetchall()
    def banned(self,uid):return self.conn.execute('SELECT 1 FROM blacklist WHERE uid=?',(uid,)).fetchone()is not None
    def ban(self,uid,r,by):self.conn.execute('INSERT OR REPLACE INTO blacklist VALUES(?,?,?)',(uid,r,by));self.conn.commit()
    def unban(self,uid):self.conn.execute('DELETE FROM blacklist WHERE uid=?',(uid,));self.conn.commit()
    def stat(self,cmd,uid,gid,ok,t):self.conn.execute('INSERT INTO stats(cmd,uid,gid,ok,t)VALUES(?,?,?,?,?)',(cmd,uid,gid,int(ok),t));self.conn.commit()
    def stats(self):
        c=self.conn
        return{"total":c.execute('SELECT COUNT(*)FROM stats').fetchone()[0],"cmds":c.execute('SELECT cmd,COUNT(*),AVG(t)FROM stats GROUP BY cmd ORDER BY COUNT(*)DESC').fetchall()}
db=Database()

# ==============================================================================
# ⏱️ RATE LIMITER & CHECKS
# ==============================================================================
class RL:
    def __init__(self):self.cd=defaultdict(lambda:defaultdict(float))
    def ok(self,uid,cmd,t=5):
        now=time.time();last=self.cd[uid][cmd]
        if now-last<t:return False,t-(now-last)
        self.cd[uid][cmd]=now;return True,0
rl=RL()

def rate(s=5):
    async def p(i:discord.Interaction)->bool:
        ok,r=rl.ok(i.user.id,i.command.name,s)
        if not ok:await i.response.send_message(f"⏳ Tunggu **{r:.1f}s**",ephemeral=True);return False
        return True
    return app_commands.check(p)

def owner():
    async def p(i:discord.Interaction)->bool:
        if i.user.id not in OWNER_IDS:await i.response.send_message("❌ Owner only!",ephemeral=True);return False
        return True
    return app_commands.check(p)

def noban():
    async def p(i:discord.Interaction)->bool:
        if db.banned(i.user.id):await i.response.send_message("🚫 Blacklisted!",ephemeral=True);return False
        return True
    return app_commands.check(p)

# ==============================================================================
# 🧠 MEMORY
# ==============================================================================
@dataclass
class M:
    role:str;content:str;ts:float
class Mem:
    def __init__(self):self.c=defaultdict(list)
    def add(self,uid,role,txt):
        now=time.time()
        self.c[uid]=[m for m in self.c[uid]if now-m.ts<1800]
        self.c[uid].append(M(role,txt,now))
        if len(self.c[uid])>10:self.c[uid]=self.c[uid][-10:]
    def get(self,uid):return[{"role":m.role,"content":m.content}for m in self.c[uid]]
    def clr(self,uid):self.c[uid]=[]
mem=Mem()

# ==============================================================================
# 🧠 AI SYSTEM
# ==============================================================================
GROQ_M=["llama-3.3-70b-versatile","llama-3.1-8b-instant"]
OPENAI_M=["gpt-4o","gpt-4o-mini"]
GEMINI_M=["gemini-2.0-flash","gemini-1.5-flash"]

def ask(prompt,sys="Kamu ahli coding.",uid=None,ctx=False):
    msgs=[{"role":"system","content":sys}]
    if ctx and uid:msgs.extend(mem.get(uid))
    msgs.append({"role":"user","content":prompt})
    
    cl=get_groq()
    if cl:
        for m in GROQ_M:
            try:
                r=cl.chat.completions.create(messages=msgs,model=m,temperature=0.7,max_tokens=4096)
                resp=r.choices[0].message.content
                if uid:mem.add(uid,"user",prompt);mem.add(uid,"assistant",resp)
                return f"⚡**[Groq]**\n{resp}",m
            except:pass
    
    cl=get_openai()
    if cl:
        for m in OPENAI_M:
            try:
                r=cl.chat.completions.create(model=m,messages=msgs,temperature=0.7,max_tokens=4096)
                resp=r.choices[0].message.content
                if uid:mem.add(uid,"user",prompt);mem.add(uid,"assistant",resp)
                return f"🤖**[OpenAI]**\n{resp}",m
            except:pass
    
    g=get_genai()
    if g:
        for m in GEMINI_M:
            try:
                sf=[{"category":c,"threshold":"BLOCK_NONE"}for c in["HARM_CATEGORY_HARASSMENT","HARM_CATEGORY_HATE_SPEECH","HARM_CATEGORY_SEXUALLY_EXPLICIT","HARM_CATEGORY_DANGEROUS_CONTENT"]]
                mdl=g.GenerativeModel(m,safety_settings=sf,system_instruction=sys)
                r=mdl.generate_content(prompt)
                if uid:mem.add(uid,"user",prompt);mem.add(uid,"assistant",r.text)
                return f"🧠**[Gemini]**\n{r.text}",m
            except:pass
    
    try:
        req=get_requests()
        for pm in["openai","mistral"]:
            r=req.get(f"https://text.pollinations.ai/{quote(sys+' '+prompt)}?model={pm}",timeout=45)
            if r.ok and len(r.text)>10:
                if uid:mem.add(uid,"user",prompt);mem.add(uid,"assistant",r.text)
                return f"🌺**[Poll]**\n{r.text}",pm
    except:pass
    return "❌ AI unavailable.","none"

def split(t,lim=1900):
    if len(t)<=lim:return[t]
    ch=[];cur=""
    for l in t.split('\n'):
        if len(cur)+len(l)+1>lim:
            if cur:ch.append(cur)
            cur=l
        else:cur+=('\n'if cur else'')+l
    if cur:ch.append(cur)
    return ch or[t[:lim]]

# ==============================================================================
# 📊 EXCEL SYSTEM
# ==============================================================================
EXCEL_SYSTEM_PROMPT = """Kamu adalah Excel Expert & Data Analyst profesional.

KEMAMPUAN:
1. RUMUS EXCEL: SUM, AVERAGE, COUNT, IF, VLOOKUP, HLOOKUP, INDEX, MATCH, SUMIF, COUNTIF, SUMIFS, COUNTIFS, IFERROR, LEFT, RIGHT, MID, LEN, TRIM, CONCATENATE, TEXT, DATE, YEAR, MONTH, DAY, NOW, TODAY, DATEDIF, NETWORKDAYS, EOMONTH, PMT, FV, PV, NPV, IRR, XNPV, XIRR, ROUND, ROUNDUP, ROUNDDOWN, ABS, MAX, MIN, LARGE, SMALL, RANK, PERCENTILE, QUARTILE, STDEV, VAR, CORREL, TREND, FORECAST, LINEST, GROWTH, TRANSPOSE, UNIQUE, SORT, FILTER, XLOOKUP, LET, LAMBDA, SEQUENCE, RANDARRAY, dll.

2. PIVOT TABLE & CHARTS: Cara membuat, menganalisis data

3. DATA ANALYSIS: Statistik, forecasting, regresi, korelasi

4. FINANCIAL: ROI, NPV, IRR, amortisasi, depreciation, break-even

5. TEMPLATES: Invoice, laporan keuangan, inventory, payroll, budget

INSTRUKSI:
- Jawab dalam Bahasa Indonesia
- Berikan rumus yang siap pakai
- Jelaskan step-by-step jika diminta
- Jika diminta buat file, return JSON format untuk data"""

class ExcelGenerator:
    """Generate Excel files with various templates and data"""
    
    @staticmethod
    def create_from_data(data, sheet_name="Sheet1", include_formulas=True):
        """Create Excel from list of dicts or 2D array"""
        pd = get_pandas()
        openpyxl = get_openpyxl()
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        output = io.BytesIO()
        
        if isinstance(data, list) and len(data) > 0:
            if isinstance(data[0], dict):
                df = pd.DataFrame(data)
            else:
                df = pd.DataFrame(data[1:], columns=data[0] if data else None)
        else:
            df = pd.DataFrame(data)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                if r_idx == 1:  # Header
                    cell.font = header_font
                    cell.fill = header_fill
        
        # Auto-width columns
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
        
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def create_invoice(data):
        """Create professional invoice"""
        openpyxl = get_openpyxl()
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoice"
        
        # Styles
        title_font = Font(bold=True, size=20, color="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = "INVOICE"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Company Info
        ws['A3'] = "Dari:"
        ws['A4'] = data.get('company', 'Nama Perusahaan')
        ws['A5'] = data.get('address', 'Alamat')
        
        ws['D3'] = "Kepada:"
        ws['D4'] = data.get('client', 'Nama Client')
        ws['D5'] = data.get('client_address', 'Alamat Client')
        
        ws['A7'] = f"No. Invoice: {data.get('invoice_no', 'INV-001')}"
        ws['D7'] = f"Tanggal: {data.get('date', time.strftime('%Y-%m-%d'))}"
        
        # Table Headers
        headers = ['No', 'Deskripsi', 'Qty', 'Harga', 'Diskon', 'Total']
        for idx, h in enumerate(headers, 1):
            cell = ws.cell(row=9, column=idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Items
        items = data.get('items', [{'desc': 'Item 1', 'qty': 1, 'price': 100000, 'discount': 0}])
        row = 10
        for idx, item in enumerate(items, 1):
            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=2, value=item.get('desc', '')).border = border
            ws.cell(row=row, column=3, value=item.get('qty', 1)).border = border
            ws.cell(row=row, column=4, value=item.get('price', 0)).border = border
            ws.cell(row=row, column=4).number_format = '#,##0'
            ws.cell(row=row, column=5, value=item.get('discount', 0)).border = border
            ws.cell(row=row, column=5).number_format = '#,##0'
            # Formula for total
            ws.cell(row=row, column=6, value=f"=(C{row}*D{row})-E{row}").border = border
            ws.cell(row=row, column=6).number_format = '#,##0'
            row += 1
        
        # Totals
        row += 1
        ws.cell(row=row, column=5, value="Subtotal:").font = Font(bold=True)
        ws.cell(row=row, column=6, value=f"=SUM(F10:F{row-2})").font = Font(bold=True)
        ws.cell(row=row, column=6).number_format = '#,##0'
        
        row += 1
        tax_rate = data.get('tax', 11)
        ws.cell(row=row, column=5, value=f"PPN ({tax_rate}%):")
        ws.cell(row=row, column=6, value=f"=F{row-1}*{tax_rate/100}")
        ws.cell(row=row, column=6).number_format = '#,##0'
        
        row += 1
        ws.cell(row=row, column=5, value="TOTAL:").font = Font(bold=True, size=12)
        ws.cell(row=row, column=6, value=f"=F{row-2}+F{row-1}").font = Font(bold=True, size=12)
        ws.cell(row=row, column=6).number_format = '#,##0'
        
        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 18
        
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def create_financial_report(data):
        """Create financial report template"""
        openpyxl = get_openpyxl()
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Laporan Keuangan"
        
        title_font = Font(bold=True, size=14)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
        
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = data.get('title', 'LAPORAN LABA RUGI')
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws['A2'] = f"Periode: {data.get('period', 'Januari - Desember 2025')}"
        
        # Revenue Section
        row = 4
        ws.cell(row=row, column=1, value="PENDAPATAN").font = Font(bold=True)
        row += 1
        
        revenues = data.get('revenues', [{'name': 'Penjualan', 'amount': 100000000}])
        for rev in revenues:
            ws.cell(row=row, column=2, value=rev['name'])
            ws.cell(row=row, column=4, value=rev['amount']).number_format = '#,##0'
            row += 1
        
        ws.cell(row=row, column=2, value="Total Pendapatan").font = Font(bold=True)
        ws.cell(row=row, column=4, value=f"=SUM(D5:D{row-1})").font = Font(bold=True)
        ws.cell(row=row, column=4).number_format = '#,##0'
        total_rev_row = row
        row += 2
        
        # Expenses Section
        ws.cell(row=row, column=1, value="BEBAN/BIAYA").font = Font(bold=True)
        row += 1
        expense_start = row
        
        expenses = data.get('expenses', [{'name': 'Gaji Karyawan', 'amount': 30000000}, {'name': 'Sewa', 'amount': 10000000}])
        for exp in expenses:
            ws.cell(row=row, column=2, value=exp['name'])
            ws.cell(row=row, column=4, value=exp['amount']).number_format = '#,##0'
            row += 1
        
        ws.cell(row=row, column=2, value="Total Beban").font = Font(bold=True)
        ws.cell(row=row, column=4, value=f"=SUM(D{expense_start}:D{row-1})").font = Font(bold=True)
        ws.cell(row=row, column=4).number_format = '#,##0'
        total_exp_row = row
        row += 2
        
        # Net Income
        ws.cell(row=row, column=1, value="LABA BERSIH").font = Font(bold=True, size=12)
        ws.cell(row=row, column=4, value=f"=D{total_rev_row}-D{total_exp_row}").font = Font(bold=True, size=12)
        ws.cell(row=row, column=4).number_format = '#,##0'
        
        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 5
        ws.column_dimensions['D'].width = 20
        
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def create_inventory(data):
        """Create inventory tracking sheet"""
        openpyxl = get_openpyxl()
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.formatting.rule import FormulaRule
        
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory"
        
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
        
        # Headers
        headers = ['Kode', 'Nama Barang', 'Kategori', 'Stok', 'Min Stok', 'Harga Beli', 'Harga Jual', 'Nilai Stok', 'Status']
        for idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Sample data or provided data
        items = data.get('items', [
            {'code': 'SKU001', 'name': 'Produk A', 'category': 'Elektronik', 'stock': 50, 'min_stock': 10, 'buy_price': 100000, 'sell_price': 150000},
            {'code': 'SKU002', 'name': 'Produk B', 'category': 'Pakaian', 'stock': 5, 'min_stock': 10, 'buy_price': 50000, 'sell_price': 80000},
        ])
        
        for row_idx, item in enumerate(items, 2):
            ws.cell(row=row_idx, column=1, value=item.get('code', '')).border = border
            ws.cell(row=row_idx, column=2, value=item.get('name', '')).border = border
            ws.cell(row=row_idx, column=3, value=item.get('category', '')).border = border
            ws.cell(row=row_idx, column=4, value=item.get('stock', 0)).border = border
            ws.cell(row=row_idx, column=5, value=item.get('min_stock', 10)).border = border
            ws.cell(row=row_idx, column=6, value=item.get('buy_price', 0)).border = border
            ws.cell(row=row_idx, column=6).number_format = '#,##0'
            ws.cell(row=row_idx, column=7, value=item.get('sell_price', 0)).border = border
            ws.cell(row=row_idx, column=7).number_format = '#,##0'
            # Formula: Stock Value
            ws.cell(row=row_idx, column=8, value=f"=D{row_idx}*F{row_idx}").border = border
            ws.cell(row=row_idx, column=8).number_format = '#,##0'
            # Formula: Status
            ws.cell(row=row_idx, column=9, value=f'=IF(D{row_idx}<=E{row_idx},"⚠️ Restock","✅ OK")').border = border
        
        # Summary row
        last_row = len(items) + 2
        ws.cell(row=last_row, column=7, value="Total Nilai:").font = Font(bold=True)
        ws.cell(row=last_row, column=8, value=f"=SUM(H2:H{last_row-1})").font = Font(bold=True)
        ws.cell(row=last_row, column=8).number_format = '#,##0'
        
        # Column widths
        widths = [10, 25, 15, 8, 10, 15, 15, 18, 12]
        for idx, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w
        
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def create_budget(data):
        """Create budget planning sheet"""
        openpyxl = get_openpyxl()
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import BarChart, Reference
        
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Budget"
        
        header_fill = PatternFill(start_color="7B1FA2", end_color="7B1FA2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
        
        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = data.get('title', 'BUDGET BULANAN 2025')
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Kategori', 'Budget', 'Aktual', 'Selisih', 'Status']
        for idx, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Budget items
        items = data.get('items', [
            {'category': 'Gaji', 'budget': 50000000, 'actual': 48000000},
            {'category': 'Marketing', 'budget': 10000000, 'actual': 12000000},
            {'category': 'Operasional', 'budget': 15000000, 'actual': 14000000},
            {'category': 'Utilities', 'budget': 5000000, 'actual': 5500000},
        ])
        
        for row_idx, item in enumerate(items, 4):
            ws.cell(row=row_idx, column=1, value=item['category']).border = border
            ws.cell(row=row_idx, column=2, value=item['budget']).border = border
            ws.cell(row=row_idx, column=2).number_format = '#,##0'
            ws.cell(row=row_idx, column=3, value=item['actual']).border = border
            ws.cell(row=row_idx, column=3).number_format = '#,##0'
            ws.cell(row=row_idx, column=4, value=f"=B{row_idx}-C{row_idx}").border = border
            ws.cell(row=row_idx, column=4).number_format = '#,##0'
            ws.cell(row=row_idx, column=5, value=f'=IF(D{row_idx}>=0,"✅ Under","❌ Over")').border = border
        
        # Totals
        last_row = len(items) + 4
        ws.cell(row=last_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=last_row, column=2, value=f"=SUM(B4:B{last_row-1})").font = Font(bold=True)
        ws.cell(row=last_row, column=2).number_format = '#,##0'
        ws.cell(row=last_row, column=3, value=f"=SUM(C4:C{last_row-1})").font = Font(bold=True)
        ws.cell(row=last_row, column=3).number_format = '#,##0'
        ws.cell(row=last_row, column=4, value=f"=B{last_row}-C{last_row}").font = Font(bold=True)
        ws.cell(row=last_row, column=4).number_format = '#,##0'
        
        # Add chart
        chart = BarChart()
        chart.title = "Budget vs Aktual"
        chart.type = "col"
        chart.style = 10
        
        data_ref = Reference(ws, min_col=2, max_col=3, min_row=3, max_row=last_row-1)
        cats = Reference(ws, min_col=1, min_row=4, max_row=last_row-1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        ws.add_chart(chart, "G3")
        
        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def create_payroll(data):
        """Create payroll sheet"""
        openpyxl = get_openpyxl()
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payroll"
        
        header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
        
        # Title
        ws.merge_cells('A1:J1')
        ws['A1'] = f"SLIP GAJI - {data.get('period', 'Januari 2025')}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['No', 'Nama', 'Jabatan', 'Gaji Pokok', 'Tunjangan', 'Lembur', 'Total Gaji', 'BPJS', 'PPh21', 'Gaji Bersih']
        for idx, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        # Employees
        employees = data.get('employees', [
            {'name': 'John Doe', 'position': 'Manager', 'basic': 15000000, 'allowance': 3000000, 'overtime': 500000},
            {'name': 'Jane Smith', 'position': 'Staff', 'basic': 8000000, 'allowance': 1500000, 'overtime': 300000},
        ])
        
        for row_idx, emp in enumerate(employees, 4):
            ws.cell(row=row_idx, column=1, value=row_idx-3).border = border
            ws.cell(row=row_idx, column=2, value=emp['name']).border = border
            ws.cell(row=row_idx, column=3, value=emp['position']).border = border
            ws.cell(row=row_idx, column=4, value=emp['basic']).border = border
            ws.cell(row=row_idx, column=4).number_format = '#,##0'
            ws.cell(row=row_idx, column=5, value=emp['allowance']).border = border
            ws.cell(row=row_idx, column=5).number_format = '#,##0'
            ws.cell(row=row_idx, column=6, value=emp['overtime']).border = border
            ws.cell(row=row_idx, column=6).number_format = '#,##0'
            # Total Gaji = Pokok + Tunjangan + Lembur
            ws.cell(row=row_idx, column=7, value=f"=D{row_idx}+E{row_idx}+F{row_idx}").border = border
            ws.cell(row=row_idx, column=7).number_format = '#,##0'
            # BPJS = 4% dari gaji pokok
            ws.cell(row=row_idx, column=8, value=f"=D{row_idx}*0.04").border = border
            ws.cell(row=row_idx, column=8).number_format = '#,##0'
            # PPh21 = 5% dari total gaji (simplified)
            ws.cell(row=row_idx, column=9, value=f"=G{row_idx}*0.05").border = border
            ws.cell(row=row_idx, column=9).number_format = '#,##0'
            # Gaji Bersih = Total - BPJS - PPh21
            ws.cell(row=row_idx, column=10, value=f"=G{row_idx}-H{row_idx}-I{row_idx}").border = border
            ws.cell(row=row_idx, column=10).number_format = '#,##0'
        
        # Totals
        last_row = len(employees) + 4
        ws.cell(row=last_row, column=3, value="TOTAL").font = Font(bold=True)
        for col in range(4, 11):
            ws.cell(row=last_row, column=col, value=f"=SUM({openpyxl.utils.get_column_letter(col)}4:{openpyxl.utils.get_column_letter(col)}{last_row-1})").font = Font(bold=True)
            ws.cell(row=last_row, column=col).number_format = '#,##0'
        
        # Column widths
        widths = [5, 20, 15, 15, 12, 12, 15, 12, 12, 15]
        for idx, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w
        
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def create_schedule(data):
        """Create schedule/calendar template"""
        openpyxl = get_openpyxl()
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Schedule"
        
        header_fill = PatternFill(start_color="00897B", end_color="00897B", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        weekend_fill = PatternFill(start_color="FFECB3", end_color="FFECB3", fill_type="solid")
        border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
        
        # Title
        month = data.get('month', 'Januari 2025')
        ws.merge_cells('A1:H1')
        ws['A1'] = f"JADWAL - {month}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Waktu', 'Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
        for idx, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Time slots
        times = data.get('times', ['08:00', '09:00', '10:00', '11:00', '12:00', '13:00', '14:00', '15:00', '16:00', '17:00'])
        for row_idx, t in enumerate(times, 4):
            ws.cell(row=row_idx, column=1, value=t).border = border
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center')
            for col in range(2, 9):
                ws.cell(row=row_idx, column=col).border = border
                if col >= 7:  # Weekend
                    ws.cell(row=row_idx, column=col).fill = weekend_fill
        
        # Column widths
        ws.column_dimensions['A'].width = 10
        for col in range(2, 9):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        wb.save(output)
        output.seek(0)
        return output

excel = ExcelGenerator()

# ==============================================================================
# 🔓 DEOBF & OBFUSCATOR (same as before)
# ==============================================================================
class Deobf:
    @staticmethod
    def dec(c):
        c=re.sub(r'\\(\d{1,3})',lambda m:chr(int(m.group(1)))if int(m.group(1))<256 else m.group(0),c)
        return re.sub(r'\\x([0-9a-fA-F]{2})',lambda m:chr(int(m.group(1),16)),c)
    @staticmethod
    def det(c):
        p={"Luraph":[r'Luraph'],"IronBrew2":[r'IB2'],"Moonsec":[r'MoonSec'],"PSU":[r'PSU'],"Luarmor":[r'Luarmor'],"Synapse":[r'SynapseXen'],"Prometheus":[r'Prometheus']}
        f=[n for n,ps in p.items()for pt in ps if re.search(pt,c,re.I)]
        if not f and c.count('(')>1000:f=["Heavy Obfuscation"]
        return",".join(f)or"Clean/Unknown"
    @staticmethod
    def strs(c):return[x for x in re.findall(r"'([^']*)'",c)+re.findall(r'"([^"]*)"',c)if 10<len(x)<500][:50]
deobf=Deobf()

class LuaObf:
    URL="https://luaobfuscator.com/api/obfuscator"
    def __init__(self,key):self.h={"apikey":key,"Content-Type":"application/json"}
    def obf(self,script,preset="medium"):
        presets={"light":{"MinifySigns":True,"Minify":True},"medium":{"MinifySigns":True,"Minify":True,"EncryptStrings":True},"heavy":{"MinifySigns":True,"Minify":True,"MinifyAll":True,"EncryptStrings":True,"ControlFlowFlattenV2AllBlocks":True},"max":{"MinifySigns":True,"Minify":True,"MinifyAll":True,"EncryptStrings":True,"ControlFlowFlattenV2AllBlocks":True,"MaxSecurityV2":True}}
        req=get_requests()
        try:
            r1=req.post(f"{self.URL}/newscript",headers=self.h,json={"script":script},timeout=30)
            if r1.status_code==401:return None,"Invalid API key"
            if r1.status_code==429:return None,"Rate limit"
            if not r1.text.strip():return None,"Empty response"
            try:d1=r1.json()
            except:return None,f"Invalid: {r1.text[:100]}"
            if not d1.get("sessionId"):return None,d1.get("message","No session")
            r2=req.post(f"{self.URL}/obfuscate",headers=self.h,json={"sessionId":d1["sessionId"],"options":presets.get(preset,presets["medium"])},timeout=60)
            if not r2.text.strip():return None,"Empty obfuscate"
            try:d2=r2.json()
            except:return None,f"Invalid: {r2.text[:100]}"
            if not d2.get("code"):return None,d2.get("message","No code")
            return d2["code"],None
        except Exception as e:return None,str(e)
lua_obf=LuaObf(KEY_LUAOBF)if KEY_LUAOBF else None

# ==============================================================================
# 🛡️ UTILS
# ==============================================================================
def headers():return{"User-Agent":random.choice(["Roblox/WinInet","RobloxStudio/WinInet"]),"Roblox-Place-Id":random.choice(["2753915549","155615604"])}
def valid(u):return u.startswith(("http://","https://"))and not any(x in u.lower()for x in["localhost","127.0.0.1","0.0.0.0"])

async def vision(url,prompt="Jelaskan gambar"):
    g=get_genai()
    if not g:return"❌ Gemini unavailable"
    try:
        aio=await get_aiohttp()
        async with aio.ClientSession()as s:
            async with s.get(url)as r:img=await r.read()
        m=g.GenerativeModel('gemini-2.0-flash')
        r=m.generate_content([prompt,{"mime_type":"image/png","data":base64.b64encode(img).decode()}])
        return r.text
    except Exception as e:return f"❌ {e}"

# ==============================================================================
# 📡 EVENTS
# ==============================================================================
@bot.event
async def on_ready():
    logger.info(f'🔥 {bot.user} | {len(bot.guilds)} guilds')
    await bot.change_presence(activity=discord.Activity(type=discord.ActivityType.watching,name="/help"))
    try:await bot.tree.sync();logger.info("✅ Synced")
    except Exception as e:logger.error(f"Sync: {e}")

@bot.tree.error
async def on_err(i:discord.Interaction,e:app_commands.AppCommandError):
    try:await i.response.send_message(f"❌ {str(e)[:100]}",ephemeral=True)
    except:pass

# ==============================================================================
# 🎮 BASIC COMMANDS
# ==============================================================================
@bot.tree.command(name="ping",description="🏓 Latency")
async def ping(i:discord.Interaction):
    e=discord.Embed(title="🏓 Pong!",color=0x00FF00)
    e.add_field(name="Latency",value=f"`{round(bot.latency*1000)}ms`")
    e.add_field(name="Servers",value=f"`{len(bot.guilds)}`")
    await i.response.send_message(embed=e)

@bot.tree.command(name="help",description="📚 Commands")
async def help_cmd(i:discord.Interaction):
    e=discord.Embed(title="📚 Bot Commands",color=0xFFD700)
    
    # Basic
    e.add_field(name="🔧 BASIC",value="`/ping` `/help` `/clear` `/history`",inline=False)
    
    # AI
    e.add_field(name="🤖 AI",value="`/tanya` - Tanya AI\n`/vision` `/analyze` - Gambar AI",inline=False)
    
    # Script Tools
    e.add_field(name="🔓 SCRIPT TOOLS",value="`/dump` `/explain` `/deobf` `/obfuscate` `/obf`",inline=False)
    
    # Excel
    e.add_field(name="📊 EXCEL",value="`/excel` - Tanya rumus Excel\n`/buat-excel` - Buat template Excel\n`/invoice` - Buat invoice\n`/laporan` - Laporan keuangan\n`/inventory` - Inventory sheet\n`/budget` - Budget planning\n`/payroll` - Slip gaji\n`/jadwal` - Jadwal/schedule",inline=False)
    
    e.set_footer(text="AI: Groq • OpenAI • Gemini • Pollinations")
    await i.response.send_message(embed=e)

# ==============================================================================
# 📊 EXCEL COMMANDS
# ==============================================================================
@bot.tree.command(name="excel",description="📊 Tanya rumus & formula Excel")
@app_commands.describe(pertanyaan="Pertanyaan tentang Excel")
@rate(8)
@noban()
async def excel_cmd(i:discord.Interaction,pertanyaan:str):
    await i.response.defer()
    ans,mdl=ask(pertanyaan,EXCEL_SYSTEM_PROMPT,i.user.id,True)
    ch=split(ans)
    e=discord.Embed(title="📊 Excel Expert",description=pertanyaan[:300],color=0x217346)
    e.set_footer(text="Tanya rumus, pivot table, analisis data, dll")
    await i.followup.send(embed=e,content=ch[0])
    for c in ch[1:]:await i.channel.send(c)

@bot.tree.command(name="buat-excel",description="📊 Buat file Excel dari data")
@app_commands.describe(data="Data dalam format: header1,header2|value1,value2|value3,value4",nama="Nama file")
@rate(15)
@noban()
async def buat_excel(i:discord.Interaction,data:str,nama:str="data"):
    await i.response.defer()
    try:
        rows=data.split('|')
        parsed=[]
        for row in rows:
            cells=row.split(',')
            parsed.append([c.strip()for c in cells])
        
        output=excel.create_from_data(parsed)
        e=discord.Embed(title="📊 Excel Created!",color=0x217346)
        e.add_field(name="Rows",value=f"`{len(parsed)}`")
        e.add_field(name="Columns",value=f"`{len(parsed[0]) if parsed else 0}`")
        await i.followup.send(embed=e,file=discord.File(output,f"{nama}.xlsx"))
    except Exception as ex:
        await i.followup.send(f"❌ Error: `{str(ex)[:200]}`")

@bot.tree.command(name="invoice",description="📄 Buat invoice profesional")
@app_commands.describe(
    company="Nama perusahaan",
    client="Nama client",
    items="Items: deskripsi:qty:harga (pisah dengan |)",
    invoice_no="Nomor invoice"
)
@rate(15)
@noban()
async def invoice_cmd(i:discord.Interaction,company:str,client:str,items:str,invoice_no:str="INV-001"):
    await i.response.defer()
    try:
        item_list=[]
        for item in items.split('|'):
            parts=item.strip().split(':')
            if len(parts)>=3:
                item_list.append({
                    'desc':parts[0].strip(),
                    'qty':int(parts[1].strip()),
                    'price':float(parts[2].strip()),
                    'discount':float(parts[3].strip())if len(parts)>3 else 0
                })
        
        data={
            'company':company,
            'client':client,
            'invoice_no':invoice_no,
            'items':item_list,
            'date':time.strftime('%Y-%m-%d')
        }
        
        output=excel.create_invoice(data)
        e=discord.Embed(title="📄 Invoice Created!",color=0x1F4E79)
        e.add_field(name="No",value=f"`{invoice_no}`")
        e.add_field(name="Items",value=f"`{len(item_list)}`")
        e.add_field(name="Client",value=client[:30])
        await i.followup.send(embed=e,file=discord.File(output,f"Invoice_{invoice_no}.xlsx"))
    except Exception as ex:
        await i.followup.send(f"❌ Error: `{str(ex)[:200]}`\n\n💡 Format: `deskripsi:qty:harga|deskripsi:qty:harga`")

@bot.tree.command(name="laporan",description="📈 Buat laporan keuangan")
@app_commands.describe(
    title="Judul laporan",
    period="Periode (contoh: Januari 2025)",
    pendapatan="Pendapatan: nama:jumlah (pisah |)",
    beban="Beban: nama:jumlah (pisah |)"
)
@rate(15)
@noban()
async def laporan_cmd(i:discord.Interaction,title:str,period:str,pendapatan:str,beban:str):
    await i.response.defer()
    try:
        revenues=[]
        for r in pendapatan.split('|'):
            parts=r.strip().split(':')
            if len(parts)>=2:
                revenues.append({'name':parts[0].strip(),'amount':float(parts[1].strip())})
        
        expenses=[]
        for e in beban.split('|'):
            parts=e.strip().split(':')
            if len(parts)>=2:
                expenses.append({'name':parts[0].strip(),'amount':float(parts[1].strip())})
        
        data={'title':title,'period':period,'revenues':revenues,'expenses':expenses}
        output=excel.create_financial_report(data)
        
        total_rev=sum(r['amount']for r in revenues)
        total_exp=sum(e['amount']for e in expenses)
        
        e=discord.Embed(title="📈 Laporan Keuangan",color=0x4472C4)
        e.add_field(name="📥 Pendapatan",value=f"`Rp {total_rev:,.0f}`")
        e.add_field(name="📤 Beban",value=f"`Rp {total_exp:,.0f}`")
        e.add_field(name="💰 Laba",value=f"`Rp {total_rev-total_exp:,.0f}`")
        await i.followup.send(embed=e,file=discord.File(output,f"Laporan_{period.replace(' ','_')}.xlsx"))
    except Exception as ex:
        await i.followup.send(f"❌ Error: `{str(ex)[:200]}`")

@bot.tree.command(name="inventory",description="📦 Buat sheet inventory")
@app_commands.describe(items="Items: kode:nama:kategori:stok:min:harga_beli:harga_jual (pisah |)")
@rate(15)
@noban()
async def inventory_cmd(i:discord.Interaction,items:str=None):
    await i.response.defer()
    try:
        item_list=[]
        if items:
            for item in items.split('|'):
                parts=item.strip().split(':')
                if len(parts)>=4:
                    item_list.append({
                        'code':parts[0].strip(),
                        'name':parts[1].strip(),
                        'category':parts[2].strip()if len(parts)>2 else'',
                        'stock':int(parts[3].strip()),
                        'min_stock':int(parts[4].strip())if len(parts)>4 else 10,
                        'buy_price':float(parts[5].strip())if len(parts)>5 else 0,
                        'sell_price':float(parts[6].strip())if len(parts)>6 else 0
                    })
        
        data={'items':item_list}if item_list else{}
        output=excel.create_inventory(data)
        e=discord.Embed(title="📦 Inventory Sheet",color=0x2E7D32)
        e.add_field(name="Items",value=f"`{len(item_list) if item_list else 'Template'}`")
        e.add_field(name="Features",value="Auto stock status, nilai stok")
        await i.followup.send(embed=e,file=discord.File(output,"Inventory.xlsx"))
    except Exception as ex:
        await i.followup.send(f"❌ Error: `{str(ex)[:200]}`")

@bot.tree.command(name="budget",description="💰 Buat budget planning")
@app_commands.describe(
    title="Judul budget",
    items="Items: kategori:budget:aktual (pisah |)"
)
@rate(15)
@noban()
async def budget_cmd(i:discord.Interaction,title:str="Budget Bulanan 2025",items:str=None):
    await i.response.defer()
    try:
        item_list=[]
        if items:
            for item in items.split('|'):
                parts=item.strip().split(':')
                if len(parts)>=3:
                    item_list.append({
                        'category':parts[0].strip(),
                        'budget':float(parts[1].strip()),
                        'actual':float(parts[2].strip())
                    })
        
        data={'title':title,'items':item_list}if item_list else{'title':title}
        output=excel.create_budget(data)
        e=discord.Embed(title="💰 Budget Planning",color=0x7B1FA2)
        e.add_field(name="Categories",value=f"`{len(item_list) if item_list else 'Template'}`")
        e.add_field(name="Features",value="Chart, status tracking")
        await i.followup.send(embed=e,file=discord.File(output,"Budget.xlsx"))
    except Exception as ex:
        await i.followup.send(f"❌ Error: `{str(ex)[:200]}`")

@bot.tree.command(name="payroll",description="💵 Buat slip gaji")
@app_commands.describe(
    period="Periode gaji",
    karyawan="Karyawan: nama:jabatan:gaji_pokok:tunjangan:lembur (pisah |)"
)
@rate(15)
@noban()
async def payroll_cmd(i:discord.Interaction,period:str="Januari 2025",karyawan:str=None):
    await i.response.defer()
    try:
        emp_list=[]
        if karyawan:
            for emp in karyawan.split('|'):
                parts=emp.strip().split(':')
                if len(parts)>=3:
                    emp_list.append({
                        'name':parts[0].strip(),
                        'position':parts[1].strip(),
                        'basic':float(parts[2].strip()),
                        'allowance':float(parts[3].strip())if len(parts)>3 else 0,
                        'overtime':float(parts[4].strip())if len(parts)>4 else 0
                    })
        
        data={'period':period,'employees':emp_list}if emp_list else{'period':period}
        output=excel.create_payroll(data)
        e=discord.Embed(title="💵 Payroll Sheet",color=0x1565C0)
        e.add_field(name="Period",value=period)
        e.add_field(name="Employees",value=f"`{len(emp_list) if emp_list else 'Template'}`")
        e.add_field(name="Features",value="BPJS, PPh21, Gaji Bersih")
        await i.followup.send(embed=e,file=discord.File(output,f"Payroll_{period.replace(' ','_')}.xlsx"))
    except Exception as ex:
        await i.followup.send(f"❌ Error: `{str(ex)[:200]}`")

@bot.tree.command(name="jadwal",description="📅 Buat template jadwal")
@app_commands.describe(month="Bulan",times="Waktu: jam1,jam2,jam3 (opsional)")
@rate(10)
@noban()
async def jadwal_cmd(i:discord.Interaction,month:str="Januari 2025",times:str=None):
    await i.response.defer()
    try:
        time_list=times.split(',')if times else None
        data={'month':month}
        if time_list:data['times']=[t.strip()for t in time_list]
        output=excel.create_schedule(data)
        e=discord.Embed(title="📅 Schedule Template",color=0x00897B)
        e.add_field(name="Month",value=month)
        e.add_field(name="Slots",value=f"`{len(time_list) if time_list else 10}`")
        await i.followup.send(embed=e,file=discord.File(output,f"Jadwal_{month.replace(' ','_')}.xlsx"))
    except Exception as ex:
        await i.followup.send(f"❌ Error: `{str(ex)[:200]}`")

# ==============================================================================
# 🔓 SCRIPT TOOLS
# ==============================================================================
@bot.tree.command(name="dump",description="🔓 Dump script")
@app_commands.describe(url="URL",raw="Raw mode")
@rate(10)
@noban()
async def dump(i:discord.Interaction,url:str,raw:bool=False):
    await i.response.defer()
    if not valid(url):return await i.followup.send("❌ Invalid URL!")
    try:
        curl=get_curl();req=get_requests()
        if raw or not SCRAPER_KEY:
            c=curl.get(url,impersonate="chrome120",headers=headers(),timeout=30).text
            m="Raw"
        else:
            c=req.get('http://api.scraperapi.com',params={'api_key':SCRAPER_KEY,'url':url},headers=headers(),timeout=90).text
            m="Scraper"
        ext="lua"
        if"<!DOCTYPE"in c[:500]:ext="html"
        elif c.strip().startswith(("{","[")):ext="json"
        e=discord.Embed(title=f"{'✅'if ext=='lua'else'⚠️'} Dump",color=0x00FF00 if ext=="lua"else 0xFFFF00)
        e.add_field(name="Size",value=f"`{len(c):,}B`")
        e.add_field(name="Type",value=f"`.{ext}`")
        await i.followup.send(embed=e,file=discord.File(io.BytesIO(c.encode()),f"dump.{ext}"))
    except Exception as ex:await i.followup.send(f"💀 {str(ex)[:200]}")

@bot.tree.command(name="tanya",description="🤖 Tanya AI")
@app_commands.describe(q="Pertanyaan",mode="Mode")
@app_commands.choices(mode=[app_commands.Choice(name="🎮 Roblox",value="roblox"),app_commands.Choice(name="🐍 Python",value="python"),app_commands.Choice(name="🌐 Web",value="web"),app_commands.Choice(name="💬 General",value="general")])
@rate(8)
@noban()
async def tanya(i:discord.Interaction,q:str,mode:str="general"):
    await i.response.defer()
    sp={"roblox":"Ahli Roblox/Lua.","python":"Ahli Python.","web":"Ahli Web Dev.","general":"Asisten helpful."}
    ic={"roblox":"🎮","python":"🐍","web":"🌐","general":"💬"}
    ans,mdl=ask(q,sp.get(mode,""),i.user.id,True)
    ch=split(ans)
    e=discord.Embed(title=f"{ic.get(mode,'🤖')} Q",description=q[:500],color=0x5865F2)
    await i.followup.send(embed=e,content=ch[0])
    for c in ch[1:]:await i.channel.send(c)

@bot.tree.command(name="explain",description="🔍 Analisa script")
@app_commands.describe(url="URL",detail="Detail")
@app_commands.choices(detail=[app_commands.Choice(name="📝 Ringkas",value="short"),app_commands.Choice(name="📋 Detail",value="detail"),app_commands.Choice(name="🛡️ Security",value="security")])
@rate(15)
@noban()
async def explain(i:discord.Interaction,url:str,detail:str="short"):
    await i.response.defer()
    if not valid(url):return await i.followup.send("❌ Invalid!")
    try:
        curl=get_curl()
        r=curl.get(url,impersonate="chrome120",timeout=15)
        lm={"short":4000,"detail":8000,"security":6000}
        pm={"short":"Jelaskan SINGKAT.","detail":"Analisa DETAIL.","security":"Security audit."}
        ans,_=ask(f"{pm[detail]}\n```lua\n{r.text[:lm.get(detail,4000)]}\n```","Script Analyst.")
        ch=split(ans)
        await i.followup.send(content=ch[0])
        for c in ch[1:]:await i.channel.send(c)
    except Exception as ex:await i.followup.send(f"❌ {str(ex)[:200]}")

@bot.tree.command(name="deobf",description="🔓 Deobfuscate")
@app_commands.describe(url="URL")
@rate(15)
@noban()
async def deobf_cmd(i:discord.Interaction,url:str):
    await i.response.defer()
    if not valid(url):return await i.followup.send("❌ Invalid!")
    try:
        curl=get_curl()
        c=curl.get(url,impersonate="chrome120",timeout=15).text
        ot=deobf.det(c);dc=deobf.dec(c[:15000]);st=deobf.strs(dc)
        e=discord.Embed(title="🔓 Deobf",color=0xE67E22)
        e.add_field(name="Size",value=f"`{len(c):,}`")
        e.add_field(name="Type",value=ot)
        e.add_field(name="Strings",value=f"`{len(st)}`")
        await i.followup.send(embed=e,file=discord.File(io.BytesIO(dc.encode()),"decoded.lua"))
    except Exception as ex:await i.followup.send(f"❌ {str(ex)[:200]}")

@bot.tree.command(name="obfuscate",description="🔒 Obfuscate URL")
@app_commands.describe(url="URL",preset="Level")
@app_commands.choices(preset=[app_commands.Choice(name="🟢 Light",value="light"),app_commands.Choice(name="🟡 Medium",value="medium"),app_commands.Choice(name="🟠 Heavy",value="heavy"),app_commands.Choice(name="🔴 Max",value="max")])
@rate(20)
@noban()
async def obfuscate_cmd(i:discord.Interaction,url:str,preset:str="medium"):
    if not KEY_LUAOBF:return await i.response.send_message("❌ API not set!",ephemeral=True)
    await i.response.defer()
    if not valid(url):return await i.followup.send("❌ Invalid!")
    try:
        curl=get_curl()
        script=curl.get(url,impersonate="chrome120",timeout=15).text
        if len(script)>500000:return await i.followup.send("❌ Max 500KB!")
        result,err=lua_obf.obf(script,preset)
        if err:return await i.followup.send(f"❌ {err}")
        e=discord.Embed(title="🔒 Obfuscated!",color=0x00FF00)
        e.add_field(name="Ori",value=f"`{len(script):,}`")
        e.add_field(name="Result",value=f"`{len(result):,}`")
        await i.followup.send(embed=e,file=discord.File(io.BytesIO(result.encode()),f"obf.lua"))
    except Exception as ex:await i.followup.send(f"💀 {str(ex)[:200]}")

@bot.tree.command(name="obf",description="🔒 Obfuscate file")
@app_commands.describe(file="File .lua",preset="Level")
@app_commands.choices(preset=[app_commands.Choice(name="🟢 Light",value="light"),app_commands.Choice(name="🟡 Medium",value="medium"),app_commands.Choice(name="🟠 Heavy",value="heavy"),app_commands.Choice(name="🔴 Max",value="max")])
@rate(20)
@noban()
async def obf_file(i:discord.Interaction,file:discord.Attachment,preset:str="medium"):
    if not KEY_LUAOBF:return await i.response.send_message("❌ API not set!",ephemeral=True)
    if not file.filename.endswith(('.lua','.txt')):return await i.response.send_message("❌ .lua only!",ephemeral=True)
    await i.response.defer()
    try:
        script=(await file.read()).decode('utf-8')
        result,err=lua_obf.obf(script,preset)
        if err:return await i.followup.send(f"❌ {err}")
        await i.followup.send(file=discord.File(io.BytesIO(result.encode()),f"obf_{file.filename}"))
    except Exception as ex:await i.followup.send(f"💀 {str(ex)[:200]}")

@bot.tree.command(name="vision",description="🖼️ Analisa gambar")
@app_commands.describe(url="URL",prompt="Question")
@rate(10)
@noban()
async def vision_cmd(i:discord.Interaction,url:str,prompt:str="Jelaskan gambar"):
    await i.response.defer()
    r=await vision(url,prompt)
    ch=split(r)
    await i.followup.send(content=ch[0])
    for c in ch[1:]:await i.channel.send(c)

@bot.tree.command(name="analyze",description="🔍 Analisa gambar upload")
@app_commands.describe(img="Image")
@rate(10)
@noban()
async def analyze(i:discord.Interaction,img:discord.Attachment):
    await i.response.defer()
    if not img.content_type or not img.content_type.startswith('image/'):return await i.followup.send("❌ Image only!")
    r=await vision(img.url,"Analisis gambar ini.")
    await i.followup.send(f"🖼️ {r[:1900]}")

@bot.tree.command(name="clear",description="🧹 Hapus memory")
async def clear(i:discord.Interaction):mem.clr(i.user.id);await i.response.send_message("🧹 Cleared!",ephemeral=True)

@bot.tree.command(name="history",description="📜 History")
async def hist(i:discord.Interaction,n:int=5):
    h=db.hist(i.user.id,min(n,10))
    if not h:return await i.response.send_message("📭 Empty.",ephemeral=True)
    e=discord.Embed(title="📜 History",color=0x3498DB)
    for idx,(p,r,t)in enumerate(h,1):e.add_field(name=f"{idx}. {p[:40]}...",value=f"```{r[:80]}...```",inline=False)
    await i.response.send_message(embed=e,ephemeral=True)

# ==============================================================================
# 👑 OWNER
# ==============================================================================
@bot.tree.command(name="stats",description="📊 Stats")
@owner()
async def stats(i:discord.Interaction):
    s=db.stats()
    e=discord.Embed(title="📊 Stats",color=0x3498DB)
    e.add_field(name="Total",value=f"`{s['total']:,}`")
    e.add_field(name="Servers",value=f"`{len(bot.guilds)}`")
    if s['cmds']:e.add_field(name="Top",value="\n".join([f"• `{c[0]}`: {c[1]}x"for c in s['cmds'][:5]]),inline=False)
    await i.response.send_message(embed=e)

@bot.tree.command(name="blacklist",description="🚫 Ban")
@owner()
async def bl(i:discord.Interaction,user:discord.User,reason:str=""):
    db.ban(user.id,reason,i.user.id);await i.response.send_message(f"🚫 {user} banned")

@bot.tree.command(name="unblacklist",description="✅ Unban")
@owner()
async def ubl(i:discord.Interaction,user:discord.User):
    db.unban(user.id);await i.response.send_message(f"✅ {user} unbanned")

@bot.tree.command(name="reload",description="🔄 Sync")
@owner()
async def reload(i:discord.Interaction):
    await i.response.defer()
    try:s=await bot.tree.sync();await i.followup.send(f"✅ {len(s)} synced!")
    except Exception as e:await i.followup.send(f"❌ {e}")

# ==============================================================================
# 🚀 START
# ==============================================================================
if __name__=="__main__":
    keep_alive()
    time.sleep(1)
    print(f"🚀 Starting | Groq{'✅'if KEY_GROQ else'❌'} OpenAI{'✅'if KEY_OPENAI else'❌'} Gemini{'✅'if KEY_GEMINI else'❌'} LuaObf{'✅'if KEY_LUAOBF else'❌'}")
    try:
        bot.run(DISCORD_TOKEN,log_handler=None)
    except discord.LoginFailure:
        print("❌ Invalid Discord Token!")
    except Exception as e:
        print(f"❌ Error: {e}")
